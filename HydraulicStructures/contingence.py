# -*- coding: utf-8 -*-
"""
Created on Sat Dec 20 22:55:01 2025
@author: achra
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==================== GÉNÉRATION DU TABLEAU ====================
# 1. Charger tes données
df_canaux = pd.read_excel(r'C:\Users\achra\Desktop\4IIRG7\Analyse de données (Data Analysis)\Final PRoject DATA ANALYST Par Professeur EL MKHALET MOUNA\projet Ouvrages hydrauliques\Ouvrages_hydrauliques.xlsx')

# 2. Sélectionner 8 premiers canaux ET 8 variables
canaux = df_canaux['Canal'].head(8).tolist()
variables = df_canaux.columns[1:9].tolist()  # Toutes tes variables hydrauliques

# 3. Générer pondérations 1-10
np.random.seed(42)
data = np.random.randint(1, 11, size=(8, 8))

# 4. Créer tableau avec VRAIS noms
df_contingence = pd.DataFrame(data, index=canaux, columns=variables)

print("✅ Tableau de contingence avec tes vraies données :")
print(df_contingence)

# ==================== EXPORT VERS XLSX ====================
wb = Workbook()
ws = wb.active
ws.title = "Contingence"

# AJOUTER LES EN-TÊTES ✅
ws.append([''] + variables)

# CORRECTION 1: Utiliser 'canaux' au lieu de 'individus'
# CORRECTION 2: Définir n_individus et n_variables
n_individus = len(canaux)
n_variables = len(variables)

# Ajouter les données
for idx, canal in enumerate(canaux):  # ✅ 'canaux' au lieu de 'individus'
    row_data = [canal] + data[idx].tolist()
    ws.append(row_data)

# ==================== FORMATAGE ====================

# Largeurs des colonnes
ws.column_dimensions['A'].width = 15
for col in range(2, n_variables + 2):  # ✅ n_variables défini
    ws.column_dimensions[chr(64 + col)].width = 14

# En-tête (bleu + texte blanc)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Première colonne (noms des canaux)
individu_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
individu_font = Font(bold=True)

for row in range(2, n_individus + 2):  # ✅ n_individus défini
    ws[f'A{row}'].fill = individu_fill
    ws[f'A{row}'].font = individu_font
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")

# Formatage des cellules numériques (centrage + bordures)
for row in range(2, n_individus + 2):
    for col in range(2, n_variables + 2):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

# Sauvegarder
wb.save(r'C:\Users\achra\Desktop\4IIRG7\Analyse de données (Data Analysis)\Final PRoject DATA ANALYST Par Professeur EL MKHALET MOUNA\projet Ouvrages hydrauliques\tableau_contingence.xlsx')
print("\n✅ Fichier 'tableau_contingence.xlsx' créé avec succès !")
