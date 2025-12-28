# -*- coding: utf-8 -*-
"""
Created on Sun Dec 21 01:54 2025
@author: achra
Génération de données cyber pour 8 canaux hydrauliques (ouvrages hydrauliques)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

cyber_data = {
    "Canal 1": {
        "Alertes_SCADA": 3,      # 1-10: pondération alertes
        "Attaques_DDoS": 2,      # 1-10: pondération attaques
        "Latence_commandes": 8,  # 1-10: faible latence = haute perf
        "Risque_global": 4       # 1-10: score pondéré
    },
    "Canal 2": {
        "Alertes_SCADA": 2,
        "Attaques_DDoS": 1,
        "Latence_commandes": 9,
        "Risque_global": 2
    },
    "Canal 3": {
        "Alertes_SCADA": 8,
        "Attaques_DDoS": 7,
        "Latence_commandes": 4,
        "Risque_global": 9
    },
    "Canal 4": {
        "Alertes_SCADA": 5,
        "Attaques_DDoS": 4,
        "Latence_commandes": 6,
        "Risque_global": 7
    },
    "Canal 5": {
        "Alertes_SCADA": 2,
        "Attaques_DDoS": 1,
        "Latence_commandes": 10,
        "Risque_global": 3
    },
    "Canal 6": {
        "Alertes_SCADA": 9,
        "Attaques_DDoS": 8,
        "Latence_commandes": 3,
        "Risque_global": 10
    },
    "Canal 7": {
        "Alertes_SCADA": 1,
        "Attaques_DDoS": 1,
        "Latence_commandes": 10,
        "Risque_global": 1
    },
    "Canal 8": {
        "Alertes_SCADA": 10,
        "Attaques_DDoS": 9,
        "Latence_commandes": 2,
        "Risque_global": 10
    },
    "Canal 9":  {
        "Alertes_SCADA": 4, 
        "Attaques_DDoS": 3, 
        "Latence_commandes": 7,  
        "Risque_global": 5
        },
    "Canal 10": {"Alertes_SCADA": 6, "Attaques_DDoS": 5, "Latence_commandes": 6,  "Risque_global": 7},
    "Canal 11": {"Alertes_SCADA": 3, "Attaques_DDoS": 2, "Latence_commandes": 9,  "Risque_global": 4},
    "Canal 12": {"Alertes_SCADA": 7, "Attaques_DDoS": 6, "Latence_commandes": 5,  "Risque_global": 8},
    "Canal 13": {"Alertes_SCADA": 2, "Attaques_DDoS": 2, "Latence_commandes": 8,  "Risque_global": 3},
    "Canal 14": {"Alertes_SCADA": 8, "Attaques_DDoS": 7, "Latence_commandes": 4,  "Risque_global": 9},
    "Canal 15": {"Alertes_SCADA": 5, "Attaques_DDoS": 4, "Latence_commandes": 7,  "Risque_global": 6},
    "Canal 16": {"Alertes_SCADA": 9, "Attaques_DDoS": 8, "Latence_commandes": 3,  "Risque_global": 10},
    "Canal 17": {"Alertes_SCADA": 1, "Attaques_DDoS": 1, "Latence_commandes": 9,  "Risque_global": 2},
    "Canal 18": {"Alertes_SCADA": 6, "Attaques_DDoS": 6, "Latence_commandes": 5,  "Risque_global": 7},
    "Canal 19": {"Alertes_SCADA": 4, "Attaques_DDoS": 3, "Latence_commandes": 8,  "Risque_global": 5},
    "Canal 20": {"Alertes_SCADA": 10,"Attaques_DDoS": 9, "Latence_commandes": 2,  "Risque_global": 10},
}

df_cyber = pd.DataFrame.from_dict(cyber_data, orient="index")
df_cyber.index.name = "Canal"

print("✅ Données cyber pour 8 canaux hydrauliques :\n")
print(df_cyber)

chemin_sortie = r"C:\Users\achra\Desktop\4IIRG7\Analyse de données (Data Analysis)\Final PRoject DATA ANALYST Par Professeur EL MKHALET MOUNA\projet Ouvrages hydrauliques\donnees_cyber_8canaux.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Cyber Ouvrages Hydrauliques"

ws.append(['Canal'] + df_cyber.columns.tolist())
for canal, row in df_cyber.iterrows():
    ws.append([canal] + row.tolist())

# Formatage
ws.column_dimensions['A'].width = 18
for col in range(2, 6):
    ws.column_dimensions[chr(64 + col)].width = 22

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

wb.save(chemin_sortie)
print(f"\n✅ Fichier sauvegardé : {chemin_sortie}")
